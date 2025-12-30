"""
Скрипт для удаления строк 29-42 из основного листа шаблона Excel
"""
from pathlib import Path
from openpyxl import load_workbook
import sys

from app.config import TEMPLATE_PATH, TEMPLATE_CONFIG_FILE


def remove_rows_from_template():
    """Удалить строки 29-42 из основного листа шаблона"""
    
    if not TEMPLATE_PATH.exists():
        print(f"Ошибка: Шаблон не найден: {TEMPLATE_PATH}")
        sys.exit(1)
    
    print(f"Открываю шаблон: {TEMPLATE_PATH}")
    
    # Загружаем шаблон
    wb = load_workbook(TEMPLATE_PATH)
    
    # Загружаем конфигурацию для определения основного листа
    import json
    if TEMPLATE_CONFIG_FILE.exists():
        with open(TEMPLATE_CONFIG_FILE, "r", encoding="utf-8") as f:
            config = json.load(f)
        main_sheet_name = config.get("main_sheet", wb.sheetnames[0])
    else:
        main_sheet_name = wb.sheetnames[0]
    
    print(f"Основной лист: {main_sheet_name}")
    
    if main_sheet_name not in wb.sheetnames:
        print(f"Ошибка: Лист '{main_sheet_name}' не найден в шаблоне")
        sys.exit(1)
    
    sheet = wb[main_sheet_name]
    
    print(f"Текущее количество строк: {sheet.max_row}")
    print("Удаляю строки 29-42...")
    
    # Удаляем строки с 29 по 42 (в обратном порядке, чтобы индексы не сдвигались)
    # В openpyxl удаление строк происходит через delete_rows(start_row, amount)
    # Удаляем 14 строк (с 29 по 42 включительно)
    rows_to_delete = 42 - 29 + 1  # 14 строк
    
    # Удаляем строки (начиная с 29, удаляем 14 строк)
    sheet.delete_rows(29, rows_to_delete)
    
    print(f"Строки 29-42 удалены. Новое количество строк: {sheet.max_row}")
    
    # Сохраняем обновленный шаблон
    print("Сохраняю обновленный шаблон...")
    wb.save(TEMPLATE_PATH)
    wb.close()
    
    print("Готово! Шаблон успешно обновлен.")


if __name__ == "__main__":
    try:
        remove_rows_from_template()
    except Exception as e:
        print(f"Ошибка при удалении строк: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)





