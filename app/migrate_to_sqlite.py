"""
Утилита миграции данных из Excel/JSON в SQLite
"""
import json
from pathlib import Path
from typing import List
import openpyxl
from openpyxl import load_workbook

from app.config import CATALOG_PATH, HISTORY_FILE, NUMBERING_FILE, DATABASE_PATH
from app.database import DatabaseManager
from app.models import CatalogEntry


def migrate_catalog_from_excel(db_manager: DatabaseManager) -> int:
    """Мигрировать данные каталога из Excel в SQLite"""
    if not CATALOG_PATH.exists():
        print(f"Файл каталога не найден: {CATALOG_PATH}")
        return 0
    
    # Проверяем, есть ли уже данные
    if db_manager.has_data("catalog_entries"):
        print("Данные каталога уже существуют в БД, пропускаем миграцию")
        return 0
    
    print(f"Миграция каталога из {CATALOG_PATH}...")
    
    wb = load_workbook(CATALOG_PATH, data_only=True)
    ws = wb.active
    
    entries_count = 0
    parts_added = set()
    
    with db_manager.get_connection() as conn:
        cursor = conn.cursor()
        
        # Пропускаем заголовок (строка 1)
        for row in ws.iter_rows(min_row=2, values_only=False):
            # Проверяем, что строка не пустая
            if not row[0].value:
                continue
            
            try:
                part_code = str(row[0].value).strip()
                workshop = str(row[1].value).strip() if row[1].value else ""
                role = str(row[2].value).strip() if row[2].value else ""
                before_name = str(row[3].value).strip() if row[3].value else ""
                unit = str(row[4].value).strip() if row[4].value else ""
                norm = float(row[5].value) if row[5].value else 0.0
                comment = str(row[6].value).strip() if len(row) > 6 and row[6].value else ""
                
                # Добавляем деталь в таблицу parts, если её ещё нет
                if part_code not in parts_added:
                    cursor.execute(
                        "INSERT OR IGNORE INTO parts (code) VALUES (?)",
                        (part_code,)
                    )
                    parts_added.add(part_code)
                
                # Добавляем запись каталога
                cursor.execute("""
                    INSERT INTO catalog_entries 
                    (part_code, workshop, role, before_name, unit, norm, comment)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                """, (part_code, workshop, role, before_name, unit, norm, comment))
                
                entries_count += 1
            except (ValueError, IndexError) as e:
                # Пропускаем некорректные строки
                print(f"Пропущена некорректная строка: {e}")
                continue
        
        wb.close()
        conn.commit()
    
    print(f"Мигрировано записей каталога: {entries_count}")
    return entries_count


def migrate_history_from_json(db_manager: DatabaseManager) -> int:
    """Мигрировать историю замен из JSON в SQLite"""
    if not HISTORY_FILE.exists():
        print(f"Файл истории не найден: {HISTORY_FILE}")
        return 0
    
    # Проверяем, есть ли уже данные
    if db_manager.has_data("material_replacements"):
        print("Данные истории уже существуют в БД, пропускаем миграцию")
        return 0
    
    print(f"Миграция истории из {HISTORY_FILE}...")
    
    try:
        with open(HISTORY_FILE, "r", encoding="utf-8") as f:
            history_data = json.load(f)
    except (json.JSONDecodeError, FileNotFoundError):
        print("Ошибка чтения файла истории")
        return 0
    
    replacements_count = 0
    
    with db_manager.get_connection() as conn:
        cursor = conn.cursor()
        
        for key, after_names in history_data.items():
            # Ключ имеет формат: "part|workshop|role|before_name"
            parts = key.split("|")
            if len(parts) != 4:
                continue
            
            part_code, workshop, role, before_name = parts
            
            # Добавляем каждую замену
            for after_name in after_names:
                if after_name:
                    cursor.execute("""
                        INSERT INTO material_replacements
                        (part_code, workshop, role, before_name, after_name)
                        VALUES (?, ?, ?, ?, ?)
                    """, (part_code, workshop, role, before_name, after_name))
                    replacements_count += 1
        
        conn.commit()
    
    print(f"Мигрировано замен материалов: {replacements_count}")
    return replacements_count


def migrate_numbering_from_json(db_manager: DatabaseManager) -> int:
    """Мигрировать нумерацию из JSON в SQLite"""
    if not NUMBERING_FILE.exists():
        print(f"Файл нумерации не найден: {NUMBERING_FILE}")
        return 0
    
    # Проверяем, есть ли уже данные
    if db_manager.has_data("numbering"):
        print("Данные нумерации уже существуют в БД, пропускаем миграцию")
        return 0
    
    print(f"Миграция нумерации из {NUMBERING_FILE}...")
    
    try:
        with open(NUMBERING_FILE, "r", encoding="utf-8") as f:
            numbering_data = json.load(f)
    except (json.JSONDecodeError, FileNotFoundError):
        print("Ошибка чтения файла нумерации")
        return 0
    
    with db_manager.get_connection() as conn:
        cursor = conn.cursor()
        
        year = numbering_data.get("year")
        last_number = numbering_data.get("last", 0)
        
        if year:
            cursor.execute("""
                INSERT INTO numbering (year, last_number)
                VALUES (?, ?)
            """, (year, last_number))
            conn.commit()
            print(f"Мигрирована нумерация: год {year}, последний номер {last_number}")
            return 1
    
    return 0


def migrate_all() -> bool:
    """Выполнить полную миграцию всех данных"""
    print("Начало миграции данных в SQLite...")
    
    db_manager = DatabaseManager()
    
    # Инициализируем схему БД
    if not db_manager.db_path.exists():
        print("Создание схемы базы данных...")
        db_manager.initialize()
    else:
        print("База данных уже существует, проверяем схему...")
        db_manager.initialize()  # Создаст таблицы, если их нет
    
    # Мигрируем данные
    catalog_count = migrate_catalog_from_excel(db_manager)
    history_count = migrate_history_from_json(db_manager)
    numbering_count = migrate_numbering_from_json(db_manager)
    
    total = catalog_count + history_count + numbering_count
    
    if total > 0:
        print(f"\nМиграция завершена. Всего мигрировано записей: {total}")
        print(f"  - Каталог: {catalog_count}")
        print(f"  - История: {history_count}")
        print(f"  - Нумерация: {numbering_count}")
    else:
        print("\nМиграция не требуется - все данные уже в БД")
    
    return True

