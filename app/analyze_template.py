"""
Скрипт для анализа структуры Excel-шаблона "Извещение на замену материалов"
Определяет координаты ячеек для заполнения, именованные диапазоны, структуру таблиц
"""
import sys
import os
from pathlib import Path

# Добавляем корневую директорию в путь
sys.path.insert(0, str(Path(__file__).parent.parent))

import io
import sys

# Устанавливаем UTF-8 для вывода в Windows
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

try:
    import openpyxl
    from openpyxl import load_workbook
except ImportError:
    print("Установите openpyxl: pip install openpyxl")
    sys.exit(1)

def analyze_template(template_path):
    """Анализирует структуру Excel-шаблона"""
    
    print(f"Анализ шаблона: {template_path}")
    print("=" * 80)
    
    # Пробуем открыть как .xlsx
    try:
        wb = load_workbook(template_path, data_only=False, keep_vba=False)
        print("[OK] Файл успешно открыт как .xlsx")
    except Exception as e1:
        print(f"[ERROR] Не удалось открыть как .xlsx: {e1}")
        print("\nПопытка открыть как .xls (требуется xlrd)...")
        try:
            import xlrd
            wb_xls = xlrd.open_workbook(template_path)
            print(f"[OK] Файл открыт как .xls, листов: {wb_xls.nsheets}")
            for i in range(wb_xls.nsheets):
                sheet = wb_xls.sheet_by_index(i)
                print(f"  Лист {i+1}: '{sheet.name}' ({sheet.nrows} строк, {sheet.ncols} столбцов)")
            return analyze_xls_template(wb_xls)
        except ImportError:
            print("[ERROR] Установите xlrd: pip install xlrd")
            return None
        except Exception as e2:
            print(f"[ERROR] Не удалось открыть как .xls: {e2}")
            return None
    
    # Анализ .xlsx
    print(f"\nЛисты в книге: {len(wb.sheetnames)}")
    for sheet_name in wb.sheetnames:
        print(f"  - {sheet_name}")
    
    # Анализ именованных диапазонов
    print("\n" + "=" * 80)
    print("ИМЕНОВАННЫЕ ДИАПАЗОНЫ:")
    if wb.defined_names:
        for name, definition in wb.defined_names.items():
            print(f"  {name}: {definition}")
    else:
        print("  (именованные диапазоны не найдены)")
    
    # Анализ первого листа (основной)
    if wb.sheetnames:
        main_sheet = wb[wb.sheetnames[0]]
        print("\n" + "=" * 80)
        print(f"АНАЛИЗ ЛИСТА: {main_sheet.title}")
        print(f"Размер: {main_sheet.max_row} строк × {main_sheet.max_column} столбцов")
        
        # Поиск ключевых ячеек
        print("\nПоиск ключевых ячеек:")
        keywords = {
            "Извещение": None,
            "№": None,
            "Дата внедрения": None,
            "Вручено": None,
            "ПЗУ": None,
            "ЗМУ": None,
            "СУ": None,
            "ОСУ": None,
            "Изделие": None,
            "Причина": None,
            "Наименование детали": None,
            "Подлежит изменению": None,
            "Вносимые изменения": None,
        }
        
        for row_idx, row in enumerate(main_sheet.iter_rows(min_row=1, max_row=100, values_only=False), 1):
            for col_idx, cell in enumerate(row, 1):
                if cell.value:
                    cell_text = str(cell.value).strip()
                    for keyword in keywords:
                        if keyword.lower() in cell_text.lower() and keywords[keyword] is None:
                            keywords[keyword] = (row_idx, col_idx, cell_text)
                            print(f"  '{keyword}': строка {row_idx}, столбец {col_idx} ({cell.coordinate}) - '{cell_text[:50]}'")
        
        # Анализ структуры таблицы изменений
        print("\n" + "=" * 80)
        print("СТРУКТУРА ТАБЛИЦЫ ИЗМЕНЕНИЙ:")
        
        # Ищем заголовки таблицы
        table_start_row = None
        for row_idx, row in enumerate(main_sheet.iter_rows(min_row=1, max_row=50, values_only=False), 1):
            row_values = [str(cell.value).strip() if cell.value else "" for cell in row]
            row_text = " ".join(row_values).lower()
            
            if "подлежит изменению" in row_text or "вносимые изменения" in row_text:
                if table_start_row is None:
                    table_start_row = row_idx
                    print(f"Начало таблицы: строка {row_idx}")
                    print("Заголовки:")
                    for col_idx, cell in enumerate(row, 1):
                        if cell.value:
                            print(f"  Столбец {col_idx} ({cell.coordinate}): '{cell.value}'")
        
        # Определяем область данных таблицы
        if table_start_row:
            data_start = table_start_row + 3  # предполагаем 2-3 строки заголовков
            print(f"\nОбласть данных таблицы начинается примерно со строки {data_start}")
            
            # Показываем несколько строк данных для понимания структуры
            print("\nПримеры строк данных (первые 5 непустых):")
            count = 0
            for row_idx in range(data_start, min(data_start + 30, main_sheet.max_row + 1)):
                row = main_sheet[row_idx]
                row_values = [str(cell.value).strip() if cell.value else "" for cell in row[:15]]
                if any(v for v in row_values):
                    count += 1
                    print(f"  Строка {row_idx}: {row_values[:10]}")
                    if count >= 5:
                        break
    
    # Анализ дополнительных листов
    if len(wb.sheetnames) > 1:
        print("\n" + "=" * 80)
        print("ДОПОЛНИТЕЛЬНЫЕ ЛИСТЫ:")
        for sheet_name in wb.sheetnames[1:]:
            sheet = wb[sheet_name]
            print(f"\nЛист: {sheet_name}")
            print(f"  Размер: {sheet.max_row} строк × {sheet.max_column} столбцов")
    
    wb.close()
    return wb

def analyze_xls_template(wb_xls):
    """Анализ старого формата .xls"""
    print("\n" + "=" * 80)
    print("АНАЛИЗ .XLS ФАЙЛА:")
    
    for i in range(wb_xls.nsheets):
        sheet = wb_xls.sheet_by_index(i)
        print(f"\nЛист {i+1}: '{sheet.name}'")
        print(f"  Размер: {sheet.nrows} строк × {sheet.ncols} столбцов")
        
        # Показываем первые 20 строк для понимания структуры
        print("  Первые строки:")
        for row_idx in range(min(20, sheet.nrows)):
            row_values = [sheet.cell_value(row_idx, col_idx) for col_idx in range(min(15, sheet.ncols))]
            row_text = " | ".join([str(v)[:30] for v in row_values if v])
            if row_text.strip():
                print(f"    Строка {row_idx+1}: {row_text[:100]}")
    
    return wb_xls

if __name__ == "__main__":
    template_path = Path(__file__).parent.parent / "Образец.xls"
    
    if not template_path.exists():
        print(f"Файл не найден: {template_path}")
        sys.exit(1)
    
    analyze_template(template_path)

