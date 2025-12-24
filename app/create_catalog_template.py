"""
Создание шаблона справочника каталога деталей и материалов
"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def create_catalog_template():
    """Создаёт шаблон справочника каталога"""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Каталог"
    
    # Заголовки
    headers = [
        "Деталь (Part)",
        "Цех (Workshop)",
        "Тип позиции (Role)",
        "Наименование материала 'до' (BeforeName)",
        "Ед. изм. (Unit)",
        "Норма (Norm)",
        "Примечание (Comment)"
    ]
    
    # Стили для заголовков
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Записываем заголовки
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(1, col_idx, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    # Примеры данных
    examples = [
        ["КИР 03.614", "ПЗУ", "краска", "Грунт-эмаль PentriProtect PUR 700 PANTONE19-4055 c", "кг", "0.0530", ""],
        ["КИР 03.614", "ПЗУ", "разбавитель", "Разбавитель PentriSolv PUR 700.3 AIR", "кг", "0.0110", ""],
        ["КИР 03.614", "ПЗУ", "отвердитель", "Отвердитель PentriHard PUR 700.3/1", "кг", "0.0040", ""],
        ["Н 026.016.02", "ЗМУ", "краска", "Грунт-эмаль PentriProtect PUR 700 PANTONE19-4055 c", "кг", "0.0260", ""],
        ["Н 026.016.02", "ЗМУ", "разбавитель", "Разбавитель PentriSolv PUR 700.3 AIR", "кг", "0.0050", ""],
        ["Н 026.016.02", "ЗМУ", "отвердитель", "Отвердитель PentriHard PUR 700.3/1", "кг", "0.0020", ""],
    ]
    
    # Записываем примеры
    for row_idx, row_data in enumerate(examples, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row_idx, col_idx, value)
            cell.border = border
            if col_idx in [5, 6]:  # Норма - числовой формат
                try:
                    cell.value = float(value) if value else None
                    cell.number_format = "0.0000"
                except:
                    pass
    
    # Настройка ширины колонок
    ws.column_dimensions['A'].width = 20  # Деталь
    ws.column_dimensions['B'].width = 10  # Цех
    ws.column_dimensions['C'].width = 15  # Тип позиции
    ws.column_dimensions['D'].width = 50  # Наименование материала
    ws.column_dimensions['E'].width = 10  # Ед. изм.
    ws.column_dimensions['F'].width = 12  # Норма
    ws.column_dimensions['G'].width = 30  # Примечание
    
    # Замораживаем первую строку
    ws.freeze_panes = "A2"
    
    # Сохраняем
    catalog_path = Path(__file__).parent.parent / "catalog" / "catalog.xlsx"
    catalog_path.parent.mkdir(exist_ok=True, parents=True)
    wb.save(catalog_path)
    
    print(f"[OK] Шаблон справочника создан: {catalog_path}")
    print("\nСтруктура справочника:")
    print("  - Деталь (Part): код детали, например 'КИР 03.614'")
    print("  - Цех (Workshop): ПЗУ, ЗМУ, СУ, ОСУ")
    print("  - Тип позиции (Role): краска, разбавитель, отвердитель, лист, круг и т.д.")
    print("  - Наименование материала 'до' (BeforeName): полное название материала")
    print("  - Ед. изм. (Unit): кг, шт, м и т.д.")
    print("  - Норма (Norm): числовое значение нормы")
    print("  - Примечание (Comment): опциональное поле")
    
    return catalog_path

if __name__ == "__main__":
    create_catalog_template()

