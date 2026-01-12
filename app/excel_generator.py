"""
Генератор Excel документов по шаблону
"""
from pathlib import Path
from datetime import date
import json
import logging
from typing import Optional, Tuple, List, Dict, Set
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from app.config import TEMPLATE_PATH, TEMPLATE_CONFIG_FILE
from app.models import DocumentData, PartChanges
from app.numbering import NumberingManager

logger = logging.getLogger(__name__)


def get_merged_cell_value(sheet, row, col):
    """
    Получить ячейку для записи, учитывая объединенные ячейки.
    Если ячейка является частью объединенной области, возвращает главную ячейку.
    """
    cell = sheet.cell(row, col)
    
    # Проверяем, является ли ячейка частью объединенной области
    for merged_range in sheet.merged_cells.ranges:
        # Проверяем, находится ли ячейка в диапазоне объединенных ячеек
        if (merged_range.min_row <= row <= merged_range.max_row and
            merged_range.min_col <= col <= merged_range.max_col):
            # Возвращаем главную ячейку (верхняя левая ячейка объединенной области)
            return sheet.cell(merged_range.min_row, merged_range.min_col)
    
    return cell


class ExcelGenerator:
    """Генератор Excel документов"""
    
    def __init__(self):
        self.template_path = TEMPLATE_PATH
        self.config = self.load_config()
        self.numbering = NumberingManager()

    # -----------------------------
    # Pagination / bounds helpers
    # -----------------------------
    def _find_table_header_row(self, sheet, max_scan_rows: int = 120) -> Optional[int]:
        """Найти строку заголовка таблицы изменений по ключевым словам."""
        # В шаблоне заголовки находятся в первых строках, поэтому скан ограничен.
        for row in range(1, min(max_scan_rows, sheet.max_row) + 1):
            parts: List[str] = []
            for col in range(1, 16):  # A..O (с запасом)
                cell = get_merged_cell_value(sheet, row, col)
                if cell.value:
                    parts.append(str(cell.value).strip())
            joined = " ".join(parts).lower()
            if "подлежит изменению" in joined or "вносимые изменения" in joined:
                return row
        return None

    def _paper_height_inches(self, sheet) -> float:
        """Высота бумаги в дюймах (нужно для расчёта вместимости по высоте)."""
        # openpyxl: paperSize=9 => A4
        paper_size = getattr(sheet.page_setup, "paperSize", None)
        orientation = getattr(sheet.page_setup, "orientation", "portrait") or "portrait"

        # A4: 8.27 x 11.69
        if paper_size == 9:
            return 11.69 if orientation == "portrait" else 8.27

        # Fallback: A4 portrait
        return 11.69 if orientation == "portrait" else 8.27

    def _page_capacity_points(self, sheet) -> float:
        """Сколько 'пунктов' (pt) по высоте доступно на странице с учётом полей и масштаба."""
        pts_per_inch = 72.0
        page_height_pts = self._paper_height_inches(sheet) * pts_per_inch

        margins = sheet.page_margins
        top = (getattr(margins, "top", 0) or 0) * pts_per_inch
        bottom = (getattr(margins, "bottom", 0) or 0) * pts_per_inch

        # ВАЖНО: header/footer в Excel — это отступы до области колонтитулов,
        # а не "дополнительные поля". Их вычитание часто даёт двойной учёт.
        printable_pts = page_height_pts - top - bottom

        scale = getattr(sheet.page_setup, "scale", None) or 100
        # scale задаётся в процентах
        return printable_pts / (scale / 100.0)

    def _row_height_points(self, sheet, row: int) -> float:
        """Высота строки в пунктах (pt).

        ВАЖНО: не учитываем признак hidden как 0, потому что мы скрываем строки
        ниже области печати на доп. листах и при этом должны сохранять
        стабильный расчёт bounds (иначе при копировании листа 'поплывёт' data_end).
        """
        rd = sheet.row_dimensions.get(row)
        if rd is not None and rd.height is not None:
            return float(rd.height)
        default_h = getattr(sheet.sheet_format, "defaultRowHeight", None)
        return float(default_h) if default_h is not None else 15.0

    def _set_rows_hidden(self, sheet, start_row: int, end_row: int, hidden: bool):
        """Скрыть/показать диапазон строк (в output-документе) без изменения стилей."""
        if end_row < start_row:
            return
        for r in range(start_row, end_row + 1):
            sheet.row_dimensions[r].hidden = hidden

    def _normalize_additional_sheet_layout(self, sheet, base_sheet):
        """Нормализовать настройки печати/разметки листа под эталонный лист 1+.

        Копируем настройки по полям (attribute-by-attribute), чтобы:
        - все доп. листы печатались одинаково (масштаб/поля/ориентация/бумага)
        - не зависеть от состояния/косяков конкретного листа шаблона (2+,3+,...)
        """
        # page_setup
        base_ps = base_sheet.page_setup
        ps = sheet.page_setup
        for attr in ("scale", "fitToWidth", "fitToHeight", "orientation", "paperSize"):
            try:
                setattr(ps, attr, getattr(base_ps, attr))
            except Exception:
                pass

        # page_margins
        base_pm = base_sheet.page_margins
        pm = sheet.page_margins
        for attr in ("left", "right", "top", "bottom", "header", "footer"):
            try:
                setattr(pm, attr, getattr(base_pm, attr))
            except Exception:
                pass

        # print_options (центровка/сетки/заголовки)
        base_po = base_sheet.print_options
        po = sheet.print_options
        for attr in ("gridLines", "headings", "horizontalCentered", "verticalCentered"):
            try:
                setattr(po, attr, getattr(base_po, attr))
            except Exception:
                pass

        # Ширины колонок: копируем то, что есть у эталона (обычно A..N)
        for col_letter, dim in base_sheet.column_dimensions.items():
            if dim is None:
                continue
            try:
                sheet.column_dimensions[col_letter].width = dim.width
            except Exception:
                pass

        logger.debug(
            f"[add layout] sheet='{sheet.title}' normalized to '{base_sheet.title}': "
            f"scale={getattr(sheet.page_setup,'scale',None)}, "
            f"paperSize={getattr(sheet.page_setup,'paperSize',None)}, "
            f"orientation={getattr(sheet.page_setup,'orientation',None)}, "
            f"margins={sheet.page_margins.left},{sheet.page_margins.right},"
            f"{sheet.page_margins.top},{sheet.page_margins.bottom}"
        )

    def _get_main_sheet_bounds(self, sheet) -> Tuple[int, int, int]:
        """Границы таблицы на основном листе: (data_start_row, data_end_row, table_physical_end_row)."""
        header_row = self._find_table_header_row(sheet)
        # На основном листе ожидаем заголовок в районе 13 строки; fallback на старое поведение.
        if header_row is None:
            data_start_row = 16
        else:
            data_start_row = header_row + 3

        # Физический конец таблицы перед нижним блоком (сейчас это 42/43 в шаблоне)
        bottom_block_start_row = 43
        table_physical_end_row = bottom_block_start_row - 1

        # Рассчитываем, сколько строк таблицы можно показать, чтобы весь лист (включая низ) остался на 1 странице.
        cap = self._page_capacity_points(sheet)
        form_end_row = sheet.max_row

        fixed_height = 0.0
        # Всё, что выше данных
        for r in range(1, data_start_row):
            fixed_height += self._row_height_points(sheet, r)
        # Нижний блок (включая подписи) всегда печатается
        for r in range(bottom_block_start_row, form_end_row + 1):
            fixed_height += self._row_height_points(sheet, r)

        max_fit_row = data_start_row - 1
        acc = fixed_height
        for r in range(data_start_row, table_physical_end_row + 1):
            acc += self._row_height_points(sheet, r)
            if acc <= cap:
                max_fit_row = r
            else:
                break

        # Если расчёт дал слишком маленькое значение, оставим хотя бы 1 строку данных.
        data_end_row = max(max_fit_row, data_start_row)

        logger.info(
            f"[bounds main] header_row={header_row}, data_start={data_start_row}, "
            f"data_end={data_end_row}, table_end={table_physical_end_row}"
        )
        return data_start_row, data_end_row, table_physical_end_row

    def _get_additional_sheet_bounds(self, sheet) -> Tuple[int, int]:
        """Границы таблицы на доп. листе: (data_start_row, data_end_row).\n\n        data_end_row вычисляем так, чтобы лист печатался ровно на 1 страницу.\n        """
        header_row = self._find_table_header_row(sheet)
        # На доп. листах заголовок обычно около 4 строки.
        if header_row is None:
            data_start_row = 7  # безопасный fallback (ниже шапки и подзаголовков)
        else:
            data_start_row = header_row + 3

        cap = self._page_capacity_points(sheet)

        # Считаем, до какой строки вообще помещается лист при печати на 1 страницу.
        acc = 0.0
        max_fit_row = 1
        max_scan = min(max(sheet.max_row, 200), 500)
        for r in range(1, max_scan + 1):
            acc += self._row_height_points(sheet, r)
            if acc <= cap:
                max_fit_row = r
            else:
                break

        # Данные не должны выходить за пределы первой страницы
        data_end_row = max(max_fit_row, data_start_row)

        logger.info(
            f"[bounds add] header_row={header_row}, data_start={data_start_row}, data_end={data_end_row}"
        )
        return data_start_row, data_end_row
    
    def load_config(self) -> dict:
        """Загрузить конфигурацию шаблона"""
        if not TEMPLATE_CONFIG_FILE.exists():
            raise FileNotFoundError(f"Конфигурация шаблона не найдена: {TEMPLATE_CONFIG_FILE}")
        
        with open(TEMPLATE_CONFIG_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    
    def generate(self, document_data: DocumentData, output_path: Path):
        """Сгенерировать документ"""
        logger.info(f"Начало генерации документа №{document_data.document_number}")
        logger.debug(f"Путь сохранения: {output_path}")
        
        if not self.template_path.exists():
            raise FileNotFoundError(f"Шаблон не найден: {self.template_path}")
        
        # Проверяем доступность директории для записи
        output_dir = output_path.parent
        try:
            # Создаём директорию, если её нет
            output_dir.mkdir(parents=True, exist_ok=True)
            
            # Проверяем права на запись в директорию
            if not output_dir.exists():
                raise PermissionError(f"Не удалось создать директорию: {output_dir}")
            
            # Пробуем создать тестовый файл для проверки прав
            test_file = output_dir / ".write_test"
            try:
                test_file.touch()
                test_file.unlink()
            except (PermissionError, OSError) as e:
                raise PermissionError(
                    f"Нет прав на запись в директорию: {output_dir}\n"
                    f"Ошибка: {e}\n\n"
                    f"Проверьте права доступа к папке или закройте файл, если он открыт в Excel."
                )
        except PermissionError:
            raise
        except Exception as e:
            raise PermissionError(
                f"Не удалось проверить доступность директории: {output_dir}\n"
                f"Ошибка: {e}"
            )
        
        # Проверяем, не открыт ли файл (если он существует)
        # На Windows файл, открытый в Excel, обычно блокируется для записи
        if output_path.exists():
            try:
                # Пробуем открыть файл в режиме записи для проверки блокировки
                # На Windows это вызовет PermissionError, если файл открыт
                with open(output_path, 'r+b') as f:
                    # Пробуем прочитать первый байт, чтобы убедиться, что файл не заблокирован
                    f.read(1)
                    f.seek(0)
            except PermissionError:
                raise PermissionError(
                    f"Файл открыт в другой программе (возможно, в Excel):\n{output_path}\n\n"
                    f"Закройте файл и попробуйте снова."
                )
            except (OSError, IOError) as e:
                # На Windows может быть ошибка доступа, если файл открыт
                if "Permission denied" in str(e) or "being used by another process" in str(e):
                    raise PermissionError(
                        f"Файл открыт в другой программе (возможно, в Excel):\n{output_path}\n\n"
                        f"Закройте файл и попробуйте снова."
                    )
                # Если другая ошибка, всё равно пытаемся сохранить
                pass
            except Exception:
                # Если другая ошибка, всё равно пытаемся сохранить
                pass
        
        # Загружаем шаблон
        wb = load_workbook(self.template_path)
        main_sheet = wb[self.config["main_sheet"]]
        logger.debug(f"Шаблон загружен. Листы в шаблоне: {wb.sheetnames}")
        
        # Получаем номер документа
        # Используем год из даты внедрения для определения номера
        impl_date = document_data.implementation_date
        year = impl_date.year if impl_date else None
        
        if not document_data.document_number:
            document_data.document_number = self.numbering.get_next_number(year)
            logger.info(f"Автоматически присвоен номер документа: {document_data.document_number}")
        else:
            # Сохраняем вручную установленный номер как использованный
            self.numbering.mark_number_as_used(document_data.document_number, year)
            logger.info(f"Используется номер документа: {document_data.document_number}")
        
        # Подсчитываем общее количество деталей с изменениями
        total_parts = len([pc for pc in document_data.part_changes 
                          if any(m.is_changed and m.after_name for m in pc.materials)])
        logger.info(f"Всего деталей с изменениями: {total_parts}")
        
        # Очищаем старые данные из шаблона
        self.clear_template_data(main_sheet)
        
        # Заполняем шапку
        self.fill_header(main_sheet, document_data)
        
        # Заполняем блок "Вручено"
        self.fill_vruhcheno(main_sheet, document_data)
        
        # Заполняем таблицу изменений
        remaining_data = self.fill_changes_table(main_sheet, document_data)
        
        # Обрабатываем дополнительные листы, если нужно
        if remaining_data:
            logger.info(f"Обнаружены данные для дополнительных страниц: {len(remaining_data)} деталей")
            self.handle_additional_sheets(wb, remaining_data, document_data)
        else:
            logger.info("Все данные поместились на первую страницу, дополнительные страницы не требуются")
            # Убеждаемся, что все дополнительные листы скрыты, если данных нет
            additional_sheet_names = sorted([name for name in wb.sheetnames if "+" in name])
            if additional_sheet_names:
                logger.debug(f"Скрываем все дополнительные листы ({len(additional_sheet_names)} шт.), так как данных нет")
                for sheet_name in additional_sheet_names:
                    sheet = wb[sheet_name]
                    sheet.sheet_state = 'hidden'
        
        # Скрываем пустые строки между таблицей изменений и нижним блоком
        self.compact_empty_rows(main_sheet)
        
        # Сохраняем через временный файл для избежания проблем с блокировкой
        try:
            # Используем временный файл в той же директории
            temp_path = output_dir / f"~{output_path.name}.tmp"
            
            # Удаляем временный файл, если он существует
            if temp_path.exists():
                try:
                    temp_path.unlink()
                except:
                    pass
            
            # Сохраняем во временный файл
            wb.save(temp_path)
            wb.close()
            
            # Если целевой файл существует, удаляем его
            if output_path.exists():
                try:
                    output_path.unlink()
                except PermissionError:
                    # Если не удалось удалить, возможно файл открыт
                    try:
                        temp_path.unlink()  # Удаляем временный файл
                    except:
                        pass
                    raise PermissionError(
                        f"Не удалось перезаписать файл. Возможно, он открыт в Excel:\n{output_path}\n\n"
                        f"Закройте файл и попробуйте снова."
                    )
            
            # Переименовываем временный файл в целевой
            try:
                temp_path.rename(output_path)
                logger.info(f"Документ успешно сохранён: {output_path}")
            except PermissionError:
                # Если не удалось переименовать, возможно директория недоступна
                try:
                    temp_path.unlink()  # Удаляем временный файл
                except:
                    pass
                raise PermissionError(
                    f"Не удалось сохранить файл:\n{output_path}\n\n"
                    f"Проверьте права доступа к папке или закройте файл, если он открыт в Excel."
                )
            except Exception as e:
                # Другие ошибки при переименовании
                try:
                    temp_path.unlink()  # Удаляем временный файл
                except:
                    pass
                raise PermissionError(
                    f"Не удалось сохранить файл:\n{output_path}\n\n"
                    f"Ошибка: {e}"
                )
        except PermissionError:
            raise
        except Exception as e:
            # Закрываем книгу в случае ошибки
            try:
                wb.close()
            except:
                pass
            
            # Преобразуем ошибки доступа в более понятные сообщения
            if "Permission denied" in str(e) or "Errno 13" in str(e):
                raise PermissionError(
                    f"Нет прав на запись файла:\n{output_path}\n\n"
                    f"Возможные причины:\n"
                    f"1. Файл открыт в Excel или другой программе\n"
                    f"2. Нет прав на запись в папку\n"
                    f"3. Сетевая папка недоступна\n\n"
                    f"Ошибка: {e}"
                )
            else:
                raise
    
    def clear_template_data(self, sheet):
        """Очистить старые данные из шаблона перед заполнением"""
        # Очищаем область таблицы изменений (начиная со строки 16 до конца данных)
        data_start_row = 16
        max_row = sheet.max_row
        
        # Очищаем данные в области таблицы (колонки A-N, строки 16-42)
        # Нижняя часть документа (строка 43 и далее) предназначена для ручного заполнения
        for row in range(data_start_row, min(max_row + 1, 43)):
            for col in range(1, 15):  # Колонки A-N
                cell = get_merged_cell_value(sheet, row, col)
                # Очищаем только значение, сохраняя форматирование
                if cell.value is not None:
                    cell.value = None
        
        # Очищаем блок "Вручено" (строка 7)
        for col in range(1, 16):
            cell = get_merged_cell_value(sheet, 7, col)
            if cell.value and str(cell.value).strip().lower() in ['х', 'x']:
                cell.value = None
        
        # Очищаем поля шапки, которые будут заполняться
        # Дата внедрения (строка 9, колонка E)
        cell = get_merged_cell_value(sheet, 9, 5)
        if cell.value:
            cell.value = None
        
        # Срок действия (строка 9, колонки 8-14)
        cell = get_merged_cell_value(sheet, 9, 8)
        if cell.value:
            cell.value = None
        
        # Изделие (строка 10, колонка B - текст "Изделие:", колонка E - название изделия)
        cell = get_merged_cell_value(sheet, 10, 2)
        if cell.value:
            cell.value = None
        cell = get_merged_cell_value(sheet, 10, 5)
        if cell.value:
            cell.value = None
        
        # Причина (строка 11, колонка B - текст "Причина:", колонка E - причина)
        cell = get_merged_cell_value(sheet, 11, 2)
        if cell.value:
            cell.value = None
        cell = get_merged_cell_value(sheet, 11, 5)
        if cell.value:
            cell.value = None
        
        # Заключение ТКО (строки 10-11, колонки 8-14)
        cell = get_merged_cell_value(sheet, 10, 8)
        if cell.value:
            cell.value = None
    
    def fill_header(self, sheet, doc_data: DocumentData):
        """Заполнить шапку документа"""
        # Номер документа (в заголовке, строка 5)
        for col in range(1, 16):
            cell = get_merged_cell_value(sheet, 5, col)
            if cell.value and "извещение" in str(cell.value).lower():
                text = str(cell.value)
                if "№" in text:
                    import re
                    new_text = re.sub(r'№\s*\d+', f'№ {doc_data.document_number}', text)
                    cell.value = new_text
                break
        
        # Дата внедрения (строка 9, колонка E)
        if doc_data.implementation_date:
            cell = get_merged_cell_value(sheet, 9, 5)
            # Сохраняем существующий формат, если он есть, иначе устанавливаем формат даты
            if not cell.number_format or cell.number_format == 'General':
                cell.number_format = "DD.MM.YYYY"
            cell.value = doc_data.implementation_date
        
        # Срок действия (строка 9, колонки 8-14 - текст "Срок действия: [значение]")
        if doc_data.validity_period:
            cell = get_merged_cell_value(sheet, 9, 8)
            cell.value = f"Срок действия: {doc_data.validity_period}"
        
        # Изделие (строка 10, колонка B - текст "Изделие:", колонка E - названия изделий)
        if doc_data.products:
            # Записываем "Изделие:" в колонку B (колонки 1-3)
            cell = get_merged_cell_value(sheet, 10, 2)
            cell.value = "Изделие:"
            # Записываем названия изделий через запятую в колонку E (колонки 5-7)
            cell = get_merged_cell_value(sheet, 10, 5)
            cell.value = ", ".join(doc_data.products)
        
        # Причина (строка 11, колонка B - текст "Причина:", колонка E - причина)
        if doc_data.reason:
            # Записываем "Причина:" в колонку B (колонки 1-3)
            cell = get_merged_cell_value(sheet, 11, 2)
            cell.value = "Причина:"
            # Записываем причину в колонку E (колонки 5-7)
            cell = get_merged_cell_value(sheet, 11, 5)
            cell.value = doc_data.reason
        
        # Заключение ТКО (строки 10-11, колонки 8-14)
        # Заполняется вручную, поэтому оставляем только текст по умолчанию
        cell = get_merged_cell_value(sheet, 10, 8)
        cell.value = "Заключение ТКО (допускается/не допускается)"
    
    def fill_vruhcheno(self, sheet, doc_data: DocumentData):
        """Заполнить блок 'Вручено'"""
        # Получаем все цеха из документа
        workshops_in_doc = doc_data.get_all_workshops()
        
        # Из анализа шаблона: строка 6 - заголовки, строка 7 - отметки
        header_row = 6
        mark_row = 7
        
        workshop_cols = {}
        
        # Ищем заголовки цехов в строке 6
        for col in range(1, 16):
            cell = get_merged_cell_value(sheet, header_row, col)
            if cell.value:
                val = str(cell.value).strip().upper()
                # Ищем цеха (ПЗУ может быть как ПДО)
                if "ПЗУ" in val or ("ПДО" in val and "КЛАДОВАЯ" not in val):
                    workshop_cols["ПЗУ"] = col
                elif "ЗМУ" in val:
                    workshop_cols["ЗМУ"] = col
                elif "СУ" in val and "ОСУ" not in val:
                    workshop_cols["СУ"] = col
                elif "ОСУ" in val:
                    workshop_cols["ОСУ"] = col
        
        # Если не нашли через заголовки, используем конфиг
        if not workshop_cols:
            findings = self.config["findings"]
            vruhcheno = findings.get("vruhcheno_marks", {})
            
            # Маппинг из конфига
            config_mapping = {
                "ПЗУ": ["ПДО", "Кладовая ПДО"],
                "ЗМУ": ["ЗМУ"],
                "СУ": ["СУ"],
                "ОСУ": ["ОСУ"]
            }
            
            for workshop in workshops_in_doc:
                if workshop in config_mapping:
                    for alt_name in config_mapping[workshop]:
                        if alt_name in vruhcheno:
                            coord_data = vruhcheno[alt_name]
                            if isinstance(coord_data, list) and len(coord_data) >= 3:
                                coord = coord_data[2]
                                sheet[coord].value = "х"
                                break
        else:
            # Ставим отметки по найденным колонкам
            for workshop in workshops_in_doc:
                if workshop in workshop_cols:
                    col = workshop_cols[workshop]
                    cell = get_merged_cell_value(sheet, mark_row, col)
                    cell.value = "х"
    
    def fill_changes_table(self, sheet, doc_data: DocumentData):
        """Заполнить таблицу изменений на первом листе
        
        Записывает ТОЛЬКО значения, не изменяя стили, форматирование, ширину колонок и т.д.
        
        Returns:
            list: Оставшиеся данные, которые не поместились (для доп. листов)
        """
        data_start_row, data_end_row, table_physical_end_row = self._get_main_sheet_bounds(sheet)
        current_row = data_start_row
        logger.info(
            f"Начало заполнения таблицы изменений. Начальная строка: {data_start_row}, "
            f"лимит (1 страница): {data_end_row}, физический конец таблицы: {table_physical_end_row}"
        )
        
        remaining_data = []
        current_part_data = None
        parts_written = 0
        materials_written = 0
        overflow_started = False
        
        for part_change in doc_data.part_changes:
            # Пропускаем данные для доп. страниц - они обрабатываются отдельно
            if part_change.additional_page_number is not None:
                # Проверяем, есть ли изменения для этой детали перед добавлением
                before_materials = [m for m in part_change.materials if m.is_changed]
                after_materials = [m for m in part_change.materials if m.after_name and m.after_name.strip()]
                if before_materials or after_materials:
                    logger.debug(f"Деталь '{part_change.part}' помечена для доп. страницы {part_change.additional_page_number}, пропускаем на первой странице")
                    remaining_data.append(part_change)
                continue
            
            # Разделяем материалы на "до" и "после" независимо
            before_materials = [m for m in part_change.materials if m.is_changed]
            after_materials = [m for m in part_change.materials if m.after_name and m.after_name.strip()]
            max_rows = max(len(before_materials), len(after_materials))
            
            # Проверяем, есть ли изменения для этой детали
            if not before_materials and not after_materials:
                logger.debug(f"Деталь '{part_change.part}' не имеет изменённых материалов, пропускаем")
                continue

            # Если уже начали переносить на доп. листы — всё остальное идёт туда
            if overflow_started:
                remaining_data.append(part_change)
                continue
            
            # ВАЖНО: Проверяем, поместится ли деталь с её материалами ДО записи
            # Деталь занимает 1 строку (название детали) + максимальное количество материалов
            rows_needed = 1 + max_rows
            rows_available = data_end_row - current_row + 1
            
            logger.debug(f"Деталь '{part_change.part}': требуется строк {rows_needed} (материалов 'до': {len(before_materials)}, 'после': {len(after_materials)}), доступно {rows_available}, текущая строка {current_row}")
            
            # Если не помещается полностью, переносим на доп. страницу
            if rows_needed > rows_available:
                logger.info(
                    f"Деталь '{part_change.part}' не помещается на первую страницу "
                    f"(нужно {rows_needed} строк, доступно {rows_available}), переносим на доп. страницы (и всё далее)"
                )
                remaining_data.append(part_change)
                overflow_started = True
                continue
            
            # Записываем деталь в колонку A (первая строка группы)
            # Только значение, стиль не трогаем
            cell = get_merged_cell_value(sheet, current_row, 1)
            cell.value = part_change.part
            
            # Записываем деталь в колонку G (правая часть)
            cell = get_merged_cell_value(sheet, current_row, 7)
            cell.value = part_change.part
            
            logger.debug(f"Записана деталь '{part_change.part}' в строку {current_row}")
            current_row += 1
            parts_written += 1
            
            # Записываем материалы для этой детали независимо в левой и правой колонках
            for i in range(max_rows):
                # Дополнительная проверка на случай, если что-то пошло не так
                if current_row > data_end_row:
                    logger.warning(
                        f"Превышен лимит строк при записи материала для детали '{part_change.part}' "
                        f"(строка {current_row} > {data_end_row})"
                    )
                    # Сохраняем оставшиеся материалы для доп. листов
                    if not current_part_data:
                        current_part_data = PartChanges(part=part_change.part)
                        remaining_data.append(current_part_data)
                    # Добавляем оставшиеся материалы "до"
                    if i < len(before_materials):
                        current_part_data.materials.append(before_materials[i])
                    # Добавляем оставшиеся материалы "после"
                    if i < len(after_materials):
                        current_part_data.materials.append(after_materials[i])
                    continue
                
                # Левая часть: "Подлежит изменению"
                if i < len(before_materials):
                    material = before_materials[i]
                    # Колонка A: Наименование материала "до"
                    cell = get_merged_cell_value(sheet, current_row, 1)
                    cell.value = material.catalog_entry.before_name
                    
                    # Колонка D (4): Ед. изм. "до"
                    cell = get_merged_cell_value(sheet, current_row, 4)
                    cell.value = material.catalog_entry.unit
                    
                    # Колонка E (5): Норма "до"
                    cell = get_merged_cell_value(sheet, current_row, 5)
                    cell.value = material.catalog_entry.norm
                    # ВАЖНО: фиксируем формат нормы (3 знака после запятой),
                    # чтобы отображение не зависело от шаблона.
                    cell.number_format = "0.000"
                
                # Правая часть: "Вносимые изменения"
                if i < len(after_materials):
                    material = after_materials[i]
                    # Колонка G (7): Наименование материала "после"
                    cell = get_merged_cell_value(sheet, current_row, 7)
                    cell.value = material.after_name
                    
                    # Колонка I (9): Ед. изм. "после"
                    cell = get_merged_cell_value(sheet, current_row, 9)
                    if material.after_unit:
                        cell.value = material.after_unit
                    else:
                        cell.value = material.catalog_entry.unit
                    
                    # Колонка J (10): Норма "после" (если указана)
                    if material.after_norm is not None:
                        cell = get_merged_cell_value(sheet, current_row, 10)
                        cell.value = material.after_norm
                        cell.number_format = "0.000"
                
                current_row += 1
                materials_written += 1
                current_part_data = None
        
        logger.info(f"Заполнение первой страницы завершено. Записано деталей: {parts_written}, материалов: {materials_written}, использовано строк: {current_row - data_start_row}")
        logger.info(f"Данных для доп. страниц: {len(remaining_data)} деталей")
        
        return remaining_data
    
    def handle_additional_sheets(self, wb, remaining_data, doc_data: DocumentData):
        """Обработать дополнительные листы при переполнении
        
        Записывает ТОЛЬКО значения, сохраняя разметку шаблона.
        Простая логика:
        1. Сначала обрабатываем все детали БЕЗ additional_page_number по существующей логике
        2. Затем находим первый пустой лист и помещаем туда детали С additional_page_number
        """
        logger.info(f"Начало обработки дополнительных страниц. Получено данных: {len(remaining_data)} деталей")
        
        # Фильтруем данные: оставляем только детали с изменёнными материалами
        filtered_remaining_data = []
        for part_change in remaining_data:
            before_materials = [m for m in part_change.materials if m.is_changed]
            after_materials = [m for m in part_change.materials if m.after_name and m.after_name.strip()]
            if before_materials or after_materials:
                filtered_remaining_data.append(part_change)
            else:
                logger.debug(f"Деталь '{part_change.part}' отфильтрована (нет изменённых материалов)")
        
        logger.info(f"После фильтрации осталось: {len(filtered_remaining_data)} деталей с изменениями")
        
        # Разделяем детали на две группы: без additional_page_number и с additional_page_number
        parts_without_page_number = []
        parts_with_page_number = []
        for part_change in filtered_remaining_data:
            if part_change.additional_page_number is None:
                parts_without_page_number.append(part_change)
            else:
                parts_with_page_number.append(part_change)
        
        logger.info(f"[handle_additional_sheets] Деталей без additional_page_number: {len(parts_without_page_number)}")
        logger.info(f"[handle_additional_sheets] Деталей с additional_page_number: {len(parts_with_page_number)}")
        
        def additional_sheet_sort_key(name: str) -> int:
            if not name.endswith("+"):
                return 10**9
            try:
                return int(name[:-1])
            except Exception:
                return 10**9

        # Ищем шаблонные дополнительные листы (обычно "1+", "2+" и т.д.)
        additional_sheet_names = sorted([name for name in wb.sheetnames if name.endswith("+")], key=additional_sheet_sort_key)
        logger.debug(f"Найдено шаблонных дополнительных листов: {len(additional_sheet_names)} ({additional_sheet_names})")
        
        if not parts_without_page_number and not parts_with_page_number:
            # Нет данных для доп. листов - скрываем все листы с "+"
            logger.info("Нет данных для дополнительных страниц, скрываем все дополнительные листы")
            for sheet_name in additional_sheet_names:
                sheet = wb[sheet_name]
                sheet.sheet_state = 'hidden'
                logger.debug(f"Скрыт лист: {sheet_name}")
            return

        # Эталонный лист для всех доп. страниц: 1+
        base_sheet = None
        if "1+" in wb.sheetnames:
            base_sheet = wb["1+"]
        elif additional_sheet_names:
            base_sheet = wb[additional_sheet_names[0]]
        else:
            base_sheet = wb[self.config["main_sheet"]]

        logger.info(
            f"[add pages] base_sheet='{base_sheet.title}' "
            f"(scale={getattr(base_sheet.page_setup,'scale',None)}, "
            f"margins={base_sheet.page_margins.left},{base_sheet.page_margins.right},"
            f"{base_sheet.page_margins.top},{base_sheet.page_margins.bottom})"
        )

        def ensure_additional_sheet(page_num: int) -> str:
            """Гарантирует наличие листа '{page_num}+' и возвращает его имя."""
            name = f"{page_num}+"
            if name in wb.sheetnames:
                return name
            new_sheet = wb.copy_worksheet(base_sheet)
            new_sheet.title = name
            logger.info(f"Создан новый лист: {name}")
            return name
        
        def is_sheet_empty(sheet) -> bool:
            """Проверяет, пуст ли лист (нет данных в области таблицы)"""
            data_start_row, data_end_row = self._get_additional_sheet_bounds(sheet)
            for row in range(data_start_row, data_end_row + 1):
                for col in range(1, 15):  # Колонки A-N
                    cell = get_merged_cell_value(sheet, row, col)
                    if cell.value is not None and str(cell.value).strip():
                        return False
            return True
        
        def write_parts_to_sheet(sheet, parts: List[PartChanges], page_num: int) -> int:
            """Записывает детали на лист. Возвращает количество записанных деталей."""
            data_start_row, data_end_row = self._get_additional_sheet_bounds(sheet)
            current_row = data_start_row
            parts_written = 0
            materials_written = 0
            
            for part_change in parts:
                before_materials = [m for m in part_change.materials if m.is_changed]
                after_materials = [m for m in part_change.materials if m.after_name and m.after_name.strip()]
                max_rows = max(len(before_materials), len(after_materials))
                
                if not before_materials and not after_materials:
                    continue
                
                rows_needed = 1 + max_rows
                rows_available = data_end_row - current_row + 1
                
                # Если не помещается, прекращаем запись
                if rows_needed > rows_available:
                    logger.warning(
                        f"[handle_additional_sheets] Деталь '{part_change.part}' не помещается на страницу {page_num}+ "
                        f"(нужно {rows_needed}, доступно {rows_available})"
                    )
                    break
                
                logger.info(f"[handle_additional_sheets] Запись детали '{part_change.part}' на страницу {page_num}+, строка {current_row}")
                
                # Записываем деталь
                cell = get_merged_cell_value(sheet, current_row, 1)
                cell.value = part_change.part
                cell = get_merged_cell_value(sheet, current_row, 7)
                cell.value = part_change.part
                current_row += 1
                parts_written += 1
                
                # Записываем материалы
                for j in range(max_rows):
                    if j < len(before_materials):
                        material = before_materials[j]
                        cell = get_merged_cell_value(sheet, current_row, 1)
                        cell.value = material.catalog_entry.before_name
                        cell = get_merged_cell_value(sheet, current_row, 4)
                        cell.value = material.catalog_entry.unit
                        cell = get_merged_cell_value(sheet, current_row, 5)
                        cell.value = material.catalog_entry.norm
                        cell.number_format = "0.000"
                    
                    if j < len(after_materials):
                        material = after_materials[j]
                        cell = get_merged_cell_value(sheet, current_row, 7)
                        cell.value = material.after_name
                        cell = get_merged_cell_value(sheet, current_row, 9)
                        cell.value = material.after_unit if material.after_unit else material.catalog_entry.unit
                        if material.after_norm is not None:
                            cell = get_merged_cell_value(sheet, current_row, 10)
                            cell.value = material.after_norm
                            cell.number_format = "0.000"
                    
                    current_row += 1
                    materials_written += 1
            
            return parts_written

        pages_with_data: Set[int] = set()
        
        # ФАЗА 1: Обрабатываем детали БЕЗ additional_page_number по существующей логике
        if parts_without_page_number:
            logger.info(f"[handle_additional_sheets] ФАЗА 1: Обработка {len(parts_without_page_number)} деталей без additional_page_number")
            carry_over: List[PartChanges] = []
            page_num = 1
            
            while parts_without_page_number or carry_over:
                parts_for_page: List[PartChanges] = []
                
                if carry_over:
                    parts_for_page.extend(carry_over)
                    carry_over = []
                
                if parts_without_page_number:
                    parts_for_page.extend(parts_without_page_number)
                    parts_without_page_number = []
                
                if not parts_for_page:
                    break
                
                sheet_name = ensure_additional_sheet(page_num)
                sheet = wb[sheet_name]
                self._normalize_additional_sheet_layout(sheet, base_sheet)
                
                data_start_row, data_end_row = self._get_additional_sheet_bounds(sheet)
                self._set_rows_hidden(sheet, 1, data_end_row, False)
                self._set_rows_hidden(sheet, data_end_row + 1, max(sheet.max_row, data_end_row + 1), True)
                
                self.clear_additional_sheet_data(sheet)
                self.fill_additional_sheet_header(sheet, doc_data)
                
                current_row = data_start_row
                parts_on_page = 0
                
                for i, part_change in enumerate(parts_for_page):
                    before_materials = [m for m in part_change.materials if m.is_changed]
                    after_materials = [m for m in part_change.materials if m.after_name and m.after_name.strip()]
                    max_rows = max(len(before_materials), len(after_materials))
                    
                    if not before_materials and not after_materials:
                        continue
                    
                    rows_needed = 1 + max_rows
                    rows_available = data_end_row - current_row + 1
                    
                    if rows_needed > rows_available:
                        # Переносим оставшиеся детали на следующую страницу
                        carry_over = parts_for_page[i:]
                        logger.warning(
                            f"[handle_additional_sheets] Переполнение на {page_num}+: деталь '{part_change.part}' "
                            f"не помещается (нужно {rows_needed}, доступно {rows_available}), "
                            f"переносим {len(carry_over)} деталей на {page_num + 1}+"
                        )
                        break
                    
                    # Записываем деталь
                    cell = get_merged_cell_value(sheet, current_row, 1)
                    cell.value = part_change.part
                    cell = get_merged_cell_value(sheet, current_row, 7)
                    cell.value = part_change.part
                    current_row += 1
                    parts_on_page += 1
                    
                    # Записываем материалы
                    for j in range(max_rows):
                        if j < len(before_materials):
                            material = before_materials[j]
                            cell = get_merged_cell_value(sheet, current_row, 1)
                            cell.value = material.catalog_entry.before_name
                            cell = get_merged_cell_value(sheet, current_row, 4)
                            cell.value = material.catalog_entry.unit
                            cell = get_merged_cell_value(sheet, current_row, 5)
                            cell.value = material.catalog_entry.norm
                            cell.number_format = "0.000"
                        
                        if j < len(after_materials):
                            material = after_materials[j]
                            cell = get_merged_cell_value(sheet, current_row, 7)
                            cell.value = material.after_name
                            cell = get_merged_cell_value(sheet, current_row, 9)
                            cell.value = material.after_unit if material.after_unit else material.catalog_entry.unit
                            if material.after_norm is not None:
                                cell = get_merged_cell_value(sheet, current_row, 10)
                                cell.value = material.after_norm
                                cell.number_format = "0.000"
                        
                        current_row += 1
                
                if parts_on_page > 0:
                    sheet.sheet_state = 'visible'
                    pages_with_data.add(page_num)
                    logger.info(f"[handle_additional_sheets] Страница {page_num}+ заполнена: {parts_on_page} деталей")
                
                page_num += 1
        
        # ФАЗА 2: Находим первый пустой лист и помещаем туда детали С additional_page_number
        # Если детали не помещаются, создаем новые листы и продолжаем запись
        if parts_with_page_number:
            logger.info(f"[handle_additional_sheets] ФАЗА 2: Поиск первого пустого листа для {len(parts_with_page_number)} деталей с additional_page_number")
            
            # Начинаем поиск с листа 2+ (лист 1+ уже может быть занят)
            target_page = 2
            found_empty = False
            while target_page <= 100:
                sheet_name = ensure_additional_sheet(target_page)
                sheet = wb[sheet_name]
                
                if is_sheet_empty(sheet):
                    logger.info(f"[handle_additional_sheets] Найден пустой лист {target_page}+, помещаем туда детали с additional_page_number")
                    found_empty = True
                    break
                else:
                    logger.info(f"[handle_additional_sheets] Лист {target_page}+ занят, проверяем следующий")
                    target_page += 1
            
            # Если не нашли пустой лист до 100+, создаем новый лист после максимального номера
            if not found_empty:
                # Находим максимальный номер существующего листа с "+"
                max_page = 0
                for name in wb.sheetnames:
                    if name.endswith("+"):
                        try:
                            page_num = int(name[:-1])
                            max_page = max(max_page, page_num)
                        except ValueError:
                            pass
                target_page = max_page + 1
                logger.info(f"[handle_additional_sheets] Все листы до 100+ заняты, создаем новый лист {target_page}+ для деталей с additional_page_number")
            
            # Записываем детали, создавая новые листы при необходимости
            remaining_parts = parts_with_page_number.copy()
            total_parts_written = 0
            
            while remaining_parts:
                sheet_name = ensure_additional_sheet(target_page)
                sheet = wb[sheet_name]
                
                self._normalize_additional_sheet_layout(sheet, base_sheet)
                data_start_row, data_end_row = self._get_additional_sheet_bounds(sheet)
                self._set_rows_hidden(sheet, 1, data_end_row, False)
                self._set_rows_hidden(sheet, data_end_row + 1, max(sheet.max_row, data_end_row + 1), True)
                
                self.clear_additional_sheet_data(sheet)
                self.fill_additional_sheet_header(sheet, doc_data)
                
                current_row = data_start_row
                parts_on_page = 0
                parts_to_keep = []
                
                for part_change in remaining_parts:
                    before_materials = [m for m in part_change.materials if m.is_changed]
                    after_materials = [m for m in part_change.materials if m.after_name and m.after_name.strip()]
                    max_rows = max(len(before_materials), len(after_materials))
                    
                    if not before_materials and not after_materials:
                        # Пропускаем детали без материалов (они уже записаны)
                        total_parts_written += 1
                        continue
                    
                    rows_needed = 1 + max_rows
                    rows_available = data_end_row - current_row + 1
                    
                    # Если не помещается, создаем новый лист и продолжаем запись
                    if rows_needed > rows_available:
                        logger.info(
                            f"[handle_additional_sheets] Деталь '{part_change.part}' не помещается на страницу {target_page}+ "
                            f"(нужно {rows_needed}, доступно {rows_available}), создаем новый лист"
                        )
                        # Оставляем эту деталь и все последующие для следующего листа
                        idx = remaining_parts.index(part_change)
                        parts_to_keep = remaining_parts[idx:]
                        break
                    
                    logger.info(f"[handle_additional_sheets] Запись детали '{part_change.part}' на страницу {target_page}+, строка {current_row}")
                    
                    # Записываем деталь
                    cell = get_merged_cell_value(sheet, current_row, 1)
                    cell.value = part_change.part
                    cell = get_merged_cell_value(sheet, current_row, 7)
                    cell.value = part_change.part
                    current_row += 1
                    parts_on_page += 1
                    total_parts_written += 1
                    
                    # Записываем материалы
                    for j in range(max_rows):
                        if j < len(before_materials):
                            material = before_materials[j]
                            cell = get_merged_cell_value(sheet, current_row, 1)
                            cell.value = material.catalog_entry.before_name
                            cell = get_merged_cell_value(sheet, current_row, 4)
                            cell.value = material.catalog_entry.unit
                            cell = get_merged_cell_value(sheet, current_row, 5)
                            cell.value = material.catalog_entry.norm
                            cell.number_format = "0.000"
                        
                        if j < len(after_materials):
                            material = after_materials[j]
                            cell = get_merged_cell_value(sheet, current_row, 7)
                            cell.value = material.after_name
                            cell = get_merged_cell_value(sheet, current_row, 9)
                            cell.value = material.after_unit if material.after_unit else material.catalog_entry.unit
                            if material.after_norm is not None:
                                cell = get_merged_cell_value(sheet, current_row, 10)
                                cell.value = material.after_norm
                                cell.number_format = "0.000"
                        
                        current_row += 1
                
                # Обновляем список оставшихся деталей
                remaining_parts = parts_to_keep
                
                if parts_on_page > 0:
                    sheet.sheet_state = 'visible'
                    pages_with_data.add(target_page)
                    logger.info(f"[handle_additional_sheets] На страницу {target_page}+ записано {parts_on_page} деталей с additional_page_number")
                
                # Если остались детали, создаем новый лист
                if remaining_parts:
                    # Находим максимальный номер существующего листа с "+"
                    max_page = 0
                    for name in wb.sheetnames:
                        if name.endswith("+"):
                            try:
                                page_num = int(name[:-1])
                                max_page = max(max_page, page_num)
                            except ValueError:
                                pass
                    target_page = max_page + 1
                    logger.info(f"[handle_additional_sheets] Создаем новый лист {target_page}+ для оставшихся {len(remaining_parts)} деталей с additional_page_number")
            
            logger.info(f"[handle_additional_sheets] Всего записано {total_parts_written} деталей с additional_page_number")
        
        # ВАЖНО: Скрываем ВСЕ неиспользуемые листы (те, на которые не было записано данных)
        # Это включает листы, которые были в шаблоне, но остались пустыми
        hidden_count = 0
        # Пересобираем список листов с '+' (включая созданные)
        all_additional_sheet_names = sorted([name for name in wb.sheetnames if name.endswith("+")], key=additional_sheet_sort_key)
        for sheet_name in all_additional_sheet_names:
            page_num = additional_sheet_sort_key(sheet_name)
            sheet = wb[sheet_name]
            
            if page_num not in pages_with_data:
                # Лист не использовался - скрываем
                sheet.sheet_state = 'hidden'
                hidden_count += 1
                logger.info(f"Скрыт неиспользуемый лист: {sheet_name} (страница {page_num}+)")
            else:
                # Лист использовался - убеждаемся, что он видим
                if sheet.sheet_state != 'visible':
                    sheet.sheet_state = 'visible'
                    logger.debug(f"Лист {sheet_name} сделан видимым")
        
        logger.info(f"Обработка дополнительных страниц завершена. Видимых страниц: {len(pages_with_data)} ({sorted(pages_with_data) if pages_with_data else 'нет'}), скрыто: {hidden_count}")

    def fill_additional_sheet_header(self, sheet, doc_data: DocumentData):
        """Заполнить шапку на дополнительном листе.

        Требования:
        - В самой верхней строке должно быть слово 'Дополнение'
        - В строке 'к извещению ... № ... от ..' номер = номер текущего документа,
          а дата после 'от' = текущая дата (дата формирования листочка)
        """
        import re

        # 1) Верхняя строка: всегда 'Дополнение'
        top_cell = get_merged_cell_value(sheet, 1, 1)
        old_top = str(top_cell.value) if top_cell.value is not None else None
        top_cell.value = "Дополнение"

        # 2) Строка 'к извещению ... № ... от ..' — заменяем номер и дату
        today_str = date.today().strftime("%d.%m.%Y")
        header_line_updated = False

        for row in range(1, 4):
            for col in range(1, 15):  # A..N
                cell = get_merged_cell_value(sheet, row, col)
                if not cell.value:
                    continue
                text = str(cell.value)
                low = text.lower()
                if "извещ" in low and "№" in text:
                    new_text = re.sub(r"№\s*\d+", f"№ {doc_data.document_number}", text)
                    new_text = re.sub(r"от\s*\d{2}\.\d{2}\.\d{4}", f"от {today_str}", new_text)
                    if new_text != text:
                        cell.value = new_text
                        header_line_updated = True
                        logger.debug(
                            f"[add header] sheet='{sheet.title}': line updated at {cell.coordinate}: "
                            f"'{text}' -> '{new_text}'"
                        )
                    break
            if header_line_updated:
                break

        logger.debug(
            f"[add header] sheet='{sheet.title}': top='{old_top}' -> 'Дополнение', "
            f"header_line_updated={header_line_updated}"
        )
    
    def compact_empty_rows(self, sheet):
        """Скрыть пустые строки между таблицей изменений и нижним блоком
        
        Находит последнюю заполненную строку в диапазоне 16-42 (колонки A-N)
        и скрывает пустые строки между ней и строкой 42, чтобы нижний блок
        был ближе для удобства печати, не нарушая структуру и форматирование.
        """
        data_start_row = 16
        data_end_row = 42  # До нижнего блока (который начинается с 43)
        
        # Сначала сбрасываем скрытие для всех строк в диапазоне (на случай повторной генерации)
        for row in range(data_start_row, data_end_row + 1):
            if row in sheet.row_dimensions:
                sheet.row_dimensions[row].hidden = False
        
        # Находим последнюю строку с данными в диапазоне таблицы изменений
        last_used_row = data_start_row - 1
        
        for row in range(data_start_row, data_end_row + 1):
            # Проверяем, есть ли хотя бы одно значение в строке (колонки A-N)
            has_data = False
            for col in range(1, 15):  # Колонки A-N
                cell = get_merged_cell_value(sheet, row, col)
                if cell.value is not None and str(cell.value).strip():
                    has_data = True
                    break
            
            if has_data:
                last_used_row = row
        
        # Если есть пустые строки между последней заполненной и нижним блоком, скрываем их
        if last_used_row < data_end_row:
            rows_to_hide = data_end_row - last_used_row
            logger.info(f"Скрываю {rows_to_hide} пустых строк между строкой {last_used_row + 1} и нижним блоком (строка 43)")
            
            # Скрываем строки (начиная с last_used_row + 1 до data_end_row)
            for row in range(last_used_row + 1, data_end_row + 1):
                # В openpyxl обращение к row_dimensions[row] автоматически создаёт объект, если его нет
                sheet.row_dimensions[row].hidden = True
                # Также устанавливаем минимальную высоту для скрытых строк
                sheet.row_dimensions[row].height = 0
            
            logger.info(f"Пустые строки скрыты. Нижний блок остаётся на строке 43, но пустые строки не видны при печати")
        else:
            logger.debug(f"Пустые строки не обнаружены (последняя заполненная строка: {last_used_row})")
    
    def clear_additional_sheet_data(self, sheet):
        """Очистить старые данные на дополнительном листе"""
        data_start_row, data_end_row = self._get_additional_sheet_bounds(sheet)

        # Очищаем ТОЛЬКО область данных таблицы (A..N), шапку не трогаем
        for row in range(data_start_row, data_end_row + 1):
            for col in range(1, 15):  # Колонки A-N
                cell = get_merged_cell_value(sheet, row, col)
                if cell.value is not None:
                    cell.value = None
