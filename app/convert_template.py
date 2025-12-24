"""
Конвертация шаблона .xls в .xlsx для удобства работы с openpyxl
"""
import sys
from pathlib import Path
import xlrd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def convert_xls_to_xlsx(xls_path, xlsx_path):
    """Конвертирует .xls в .xlsx с сохранением структуры"""
    print(f"Конвертация {xls_path} -> {xlsx_path}")
    
    # Открываем .xls
    wb_xls = xlrd.open_workbook(xls_path)
    
    # Создаём новую книгу .xlsx
    wb_xlsx = Workbook()
    wb_xlsx.remove(wb_xlsx.active)  # Удаляем дефолтный лист
    
    # Конвертируем каждый лист
    for sheet_idx in range(wb_xls.nsheets):
        sheet_xls = wb_xls.sheet_by_index(sheet_idx)
        sheet_xlsx = wb_xlsx.create_sheet(title=sheet_xls.name)
        
        print(f"  Лист {sheet_idx + 1}: {sheet_xls.name} ({sheet_xls.nrows} строк)")
        
        # Копируем данные
        for row_idx in range(sheet_xls.nrows):
            for col_idx in range(sheet_xls.ncols):
                cell_xls = sheet_xls.cell(row_idx, col_idx)
                cell_xlsx = sheet_xlsx.cell(row_idx + 1, col_idx + 1)
                
                # Копируем значение
                if cell_xls.ctype == xlrd.XL_CELL_DATE:
                    # Обработка дат Excel
                    try:
                        from datetime import datetime
                        date_tuple = xlrd.xldate_as_tuple(cell_xls.value, wb_xls.datemode)
                        if date_tuple[0] != 0:  # Если это реальная дата
                            cell_xlsx.value = datetime(*date_tuple)
                        else:
                            cell_xlsx.value = cell_xls.value
                    except:
                        cell_xlsx.value = cell_xls.value
                elif cell_xls.ctype == xlrd.XL_CELL_NUMBER:
                    cell_xlsx.value = cell_xls.value
                elif cell_xls.ctype == xlrd.XL_CELL_TEXT:
                    cell_xlsx.value = cell_xls.value
                elif cell_xls.ctype == xlrd.XL_CELL_BOOLEAN:
                    cell_xlsx.value = cell_xls.value
                else:
                    cell_xlsx.value = cell_xls.value
        
        # Настраиваем ширину колонок (примерно)
        for col_idx in range(min(sheet_xls.ncols, 15)):
            sheet_xlsx.column_dimensions[get_column_letter(col_idx + 1)].width = 15
    
    # Сохраняем
    wb_xlsx.save(xlsx_path)
    print(f"[OK] Конвертация завершена: {xlsx_path}")
    return wb_xlsx

if __name__ == "__main__":
    xls_path = Path(__file__).parent.parent / "Образец.xls"
    xlsx_path = Path(__file__).parent.parent / "templates" / "Izveshchenie_template.xlsx"
    
    xlsx_path.parent.mkdir(exist_ok=True)
    
    if not xls_path.exists():
        print(f"Файл не найден: {xls_path}")
        sys.exit(1)
    
    convert_xls_to_xlsx(xls_path, xlsx_path)

