"""
Импортёр данных из Парсинг.xlsx в наборы материалов
"""
import shutil
from pathlib import Path
from typing import List, Dict, Optional, Tuple, Tuple as TupleType
from datetime import datetime
import openpyxl
from openpyxl import load_workbook

from app.config import PROJECT_ROOT, DATABASE_PATH
from app.models import CatalogEntry
from app.catalog_loader import CatalogLoader
from app.database import DatabaseManager
from app.product_store import ProductStore


# Материалы для правила "аналог КИР 03.614"
KIR_03_614_FROM_MATERIALS = [
    "Грунт-эмаль PentriProtect PUR 700 PANTONE19-4055 с",
    "Разбавитель PentriSolv PUR 700.3 AIR",
    "Отвердитель PentriHard PUR 700.3/1"
]

KIR_03_614_TO_MATERIALS = [
    "Грунт-эмаль ЯрЛИсоат 1861 голубая даль RAL",
    "Разбавитель ЯрЛИ 667",
    "Грунтовка ЯрЛИ ЭФ-065 красно-коричневая"
]


def backup_database() -> Optional[Path]:
    """Создать резервную копию базы данных"""
    if not DATABASE_PATH.exists():
        return None
    
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    backup_path = DATABASE_PATH.parent / f"app.db.bak-{timestamp}"
    
    try:
        shutil.copy2(DATABASE_PATH, backup_path)
        return backup_path
    except Exception as e:
        print(f"Ошибка при создании резервной копии: {e}")
        return None


def parse_parsing_file(file_path: Path) -> Dict[str, List[Tuple[str, str, float]]]:
    """
    Парсинг файла Парсинг.xlsx
    
    Возвращает словарь: {деталь: [(материал, ед.изм., норма), ...]}
    
    Колонки:
    - A: материал
    - B: ед. изм.
    - C: деталь
    - D: норма
    """
    if not file_path.exists():
        raise FileNotFoundError(f"Файл не найден: {file_path}")
    
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    
    result: Dict[str, List[Tuple[str, str, float]]] = {}
    skipped_rows = 0
    processed_rows = 0
    skipped_parts = []  # Детали, которые были пропущены
    
    # Пропускаем заголовок (строка 1), начинаем со строки 2
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        # Получаем значения с проверкой на None
        material_val = row[0].value if len(row) > 0 else None
        unit_val = row[1].value if len(row) > 1 else None
        part_val = row[2].value if len(row) > 2 else None
        norm_val = row[3].value if len(row) > 3 else None
        
        # Проверяем, что есть деталь (это обязательное поле)
        if not part_val:
            skipped_rows += 1
            continue
        
        part = str(part_val).strip()
        
        # Если деталь пустая после strip, пропускаем
        if not part:
            skipped_rows += 1
            continue
        
        # Если нет материала, но есть деталь - всё равно добавляем (материал может быть пустым)
        if not material_val:
            material = ""
        else:
            material = str(material_val).strip()
        
        try:
            unit = str(unit_val).strip() if unit_val else ""
            norm = float(norm_val) if norm_val else 0.0
            
            # Добавляем даже если материал пустой (деталь важнее)
            if part not in result:
                result[part] = []
            
            result[part].append((material, unit, norm))
            processed_rows += 1
        except (ValueError, IndexError) as e:
            # Пропускаем некорректные строки, но сохраняем информацию о детали
            skipped_rows += 1
            skipped_parts.append((part, row_num, str(e)))
            print(f"Пропущена некорректная строка {row_num} (деталь: '{part}'): {e}")
            print(f"  Значения: A={material_val}, B={unit_val}, C={part_val}, D={norm_val}")
            continue
    
    print(f"Обработано строк: {processed_rows}, пропущено: {skipped_rows}")
    if skipped_parts:
        print(f"Пропущено деталей из-за ошибок: {len(skipped_parts)}")
        # Показываем первые 10 пропущенных деталей
        for part, row_num, error in skipped_parts[:10]:
            print(f"  - Строка {row_num}, деталь '{part}': {error}")
        if len(skipped_parts) > 10:
            print(f"  ... и ещё {len(skipped_parts) - 10} деталей")
    
    # Выводим список всех найденных деталей (для отладки)
    all_parts = sorted(result.keys())
    print(f"\nНайдено уникальных деталей: {len(all_parts)}")
    
    # Ищем детали, содержащие "04.010" для проверки
    parts_04010 = [p for p in all_parts if "04.010" in p.upper()]
    if parts_04010:
        print(f"Детали, содержащие '04.010': {parts_04010}")
    else:
        print("⚠ Детали с '04.010' не найдены в файле")
        # Показываем похожие детали
        similar = [p for p in all_parts if "04" in p.upper() and "01" in p.upper()]
        if similar:
            print(f"Похожие детали (содержат '04' и '01'): {similar[:10]}")
    
    wb.close()
    return result


def normalize_material_name(name: str) -> str:
    """Нормализовать название материала для сравнения (убрать лишние пробелы, привести к верхнему регистру)"""
    return " ".join(name.upper().split())


def check_if_has_kir_materials(materials: List[Tuple[str, str, float]], part_code: str = "") -> Tuple[bool, List[str]]:
    """
    Проверить, содержит ли список материалов все три материала из правила КИР 03.614
    
    Returns:
        (bool, List[str]): (найдены ли все материалы, список найденных материалов)
    """
    material_names = [mat[0] for mat in materials]
    
    # Нормализуем названия материалов для сравнения
    normalized_material_names = [normalize_material_name(name) for name in material_names]
    normalized_required = [normalize_material_name(name) for name in KIR_03_614_FROM_MATERIALS]
    
    found_materials = []
    missing_materials = []
    
    # Проверяем каждый требуемый материал
    for i, required_material in enumerate(KIR_03_614_FROM_MATERIALS):
        normalized_required_name = normalized_required[i]
        
        # Ищем совпадение (точное или нормализованное)
        found = False
        for j, normalized_name in enumerate(normalized_material_names):
            if normalized_required_name == normalized_name:
                found = True
                found_materials.append(material_names[j])
                break
        
        if not found:
            missing_materials.append(required_material)
    
    # Логируем для отладки
    if part_code:
        if len(found_materials) > 0:
            print(f"  Деталь '{part_code}': найдено материалов КИР 03.614: {len(found_materials)}/3")
            print(f"    Найдены: {found_materials}")
            if missing_materials:
                print(f"    Отсутствуют: {missing_materials}")
                print(f"    Все материалы детали: {material_names}")
    
    # Все три материала должны быть найдены
    return len(found_materials) == 3, found_materials


def create_to_materials_for_kir_analog() -> List[CatalogEntry]:
    """Создать материалы 'после' для аналога КИР 03.614"""
    to_materials = []
    
    for material_name in KIR_03_614_TO_MATERIALS:
        entry = CatalogEntry(
            part="",  # Будет установлено позже
            workshop="",  # Пустое, как выбрал пользователь
            role="",  # Пустое
            before_name=material_name,
            unit="кг",  # Как выбрал пользователь
            norm=0.0,  # Как выбрал пользователь
            comment=""
        )
        to_materials.append(entry)
    
    return to_materials


def import_parsing_file(
    file_path: Optional[Path] = None,
    catalog_loader: Optional[CatalogLoader] = None,
    product_store: Optional[ProductStore] = None,
    db_manager: Optional[DatabaseManager] = None
) -> Tuple[int, int, List[str]]:
    """
    Импортировать данные из Парсинг.xlsx
    
    Args:
        file_path: Путь к файлу Парсинг.xlsx (если None, используется PROJECT_ROOT / "Парсинг.xlsx")
        catalog_loader: Загрузчик каталога (если None, создаётся новый)
        product_store: Хранилище изделий (если None, создаётся новое)
        db_manager: Менеджер БД (если None, создаётся новый)
    
    Returns:
        Tuple[количество_импортированных_деталей, количество_созданных_наборов, список_ошибок]
    """
    if file_path is None:
        file_path = PROJECT_ROOT / "Парсинг.xlsx"
    
    if catalog_loader is None:
        if db_manager is None:
            db_manager = DatabaseManager()
        catalog_loader = CatalogLoader(db_manager)
    
    if product_store is None:
        if db_manager is None:
            db_manager = DatabaseManager()
        product_store = ProductStore(db_manager)
    
    # Создаём резервную копию БД
    backup_path = backup_database()
    if backup_path:
        print(f"Создана резервная копия БД: {backup_path}")
    else:
        print("Предупреждение: не удалось создать резервную копию БД")
    
    # Парсим файл
    print(f"Чтение файла {file_path}...")
    parsed_data = parse_parsing_file(file_path)
    print(f"Найдено уникальных деталей в файле: {len(parsed_data)}")
    print(f"Список всех деталей из файла: {sorted(parsed_data.keys())[:10]}..." if len(parsed_data) > 10 else f"Список всех деталей: {sorted(parsed_data.keys())}")
    
    imported_parts = 0
    created_sets = 0
    skipped_existing_parts = 0  # Детали, которые были пропущены из-за существующих наборов
    errors = []
    failed_parts = []  # Детали, которые не удалось импортировать
    
    # Получаем список всех изделий для привязки
    all_products = product_store.get_all_products()
    product_ids = [pid for pid, _ in all_products]
    
    # Импортируем каждую деталь
    for part_code, materials_data in parsed_data.items():
        try:
            # Создаём материалы "до"
            from_materials = []
            for material_name, unit, norm in materials_data:
                entry = CatalogEntry(
                    part=part_code,
                    workshop="",  # Пустое, как выбрал пользователь
                    role="",  # Пустое
                    before_name=material_name,
                    unit=unit,
                    norm=norm,
                    comment=""
                )
                from_materials.append(entry)
            
            # Проверяем правило "аналог КИР 03.614"
            to_materials = []
            has_kir_materials, found_kir_materials = check_if_has_kir_materials(materials_data, part_code)
            if has_kir_materials:
                print(f"  ✓ Деталь '{part_code}': обнаружены ВСЕ материалы КИР 03.614, добавляем материалы 'после'")
                to_materials = create_to_materials_for_kir_analog()
                # Устанавливаем part_code для материалов "после"
                for mat in to_materials:
                    mat.part = part_code
            elif len(found_kir_materials) > 0:
                print(f"  ⚠ Деталь '{part_code}': найдены только {len(found_kir_materials)}/3 материалов КИР 03.614, материалы 'после' НЕ добавлены")
            
            # Проверяем, есть ли уже наборы для этой детали
            existing_sets = catalog_loader.get_replacement_sets_by_part(part_code)
            if existing_sets:
                skipped_existing_parts += 1
                print(f"  ⊘ Пропущена деталь '{part_code}': уже существует {len(existing_sets)} набор(ов)")
                continue  # Переходим к следующей детали
            
            # Создаём набор (set_name будет содержать префикс для идентификации импортированных)
            set_name = f"Импорт из Парсинг.xlsx - {datetime.now().strftime('%Y-%m-%d %H:%M')}"
            set_id = catalog_loader.add_replacement_set(
                part_code=part_code,
                from_materials=from_materials,
                to_materials=to_materials,
                set_name=set_name
            )
            
            if set_id:
                imported_parts += 1
                created_sets += 1
                print(f"  ✓ Импортирована деталь '{part_code}': {len(from_materials)} материалов 'до', {len(to_materials)} материалов 'после'")
                
                # Привязываем деталь ко всем изделиям
                for product_id in product_ids:
                    product_store.link_part_to_product(product_id, part_code)
                
                print(f"  ✓ Деталь '{part_code}' привязана ко всем изделиям ({len(product_ids)} шт.)")
            else:
                error_msg = f"Не удалось создать набор для детали '{part_code}'"
                errors.append(error_msg)
                failed_parts.append(part_code)
                print(f"  ✗ {error_msg}")
        
        except Exception as e:
            error_msg = f"Ошибка при импорте детали '{part_code}': {e}"
            errors.append(error_msg)
            failed_parts.append(part_code)
            print(f"  ✗ {error_msg}")
            import traceback
            print(f"  Детали ошибки: {traceback.format_exc()}")
    
    # Обновляем кэш каталога (это сбросит _parts_cache)
    catalog_loader.load()
    
    # Дополнительно сбрасываем кэш списка деталей на всякий случай
    catalog_loader._parts_cache = None
    
    # Проверяем, что детали действительно попали в БД
    all_parts_in_db = catalog_loader.get_all_parts()
    imported_parts_in_db = [p for p in parsed_data.keys() if p in all_parts_in_db]
    missing_parts = [p for p in parsed_data.keys() if p not in all_parts_in_db]
    
    print(f"\nИмпорт завершён:")
    print(f"  - Найдено деталей в файле: {len(parsed_data)}")
    print(f"  - Успешно импортировано деталей: {imported_parts}")
    print(f"  - Пропущено деталей (уже существуют наборы): {skipped_existing_parts}")
    print(f"  - Создано наборов: {created_sets}")
    print(f"  - Деталей в БД после импорта: {len(all_parts_in_db)}")
    print(f"  - Импортированных деталей найдено в БД: {len(imported_parts_in_db)}")
    
    if missing_parts:
        print(f"  ⚠ Деталей не найдено в БД после импорта: {len(missing_parts)}")
        print(f"    Первые 10: {missing_parts[:10]}")
        if len(missing_parts) > 10:
            print(f"    ... и ещё {len(missing_parts) - 10} деталей")
    
    if failed_parts:
        print(f"  ⚠ Деталей с ошибками импорта: {len(failed_parts)}")
        print(f"    Первые 10: {failed_parts[:10]}")
        if len(failed_parts) > 10:
            print(f"    ... и ещё {len(failed_parts) - 10} деталей")
    
    if errors:
        print(f"  - Всего ошибок: {len(errors)}")
        for error in errors[:10]:  # Показываем первые 10 ошибок
            print(f"    {error}")
        if len(errors) > 10:
            print(f"    ... и ещё {len(errors) - 10} ошибок")
    
    return imported_parts, created_sets, errors

