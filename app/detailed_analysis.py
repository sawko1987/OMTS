"""
Детальный анализ координат шаблона для заполнения
"""
import sys
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def detailed_analysis():
    """Детальный анализ координат"""
    template_path = Path(__file__).parent.parent / "templates" / "Izveshchenie_template.xlsx"
    
    if not template_path.exists():
        print(f"Шаблон не найден: {template_path}")
        return
    
    wb = load_workbook(template_path, data_only=False)
    main_sheet = wb[wb.sheetnames[0]]
    
    print("=" * 80)
    print("ДЕТАЛЬНЫЙ АНАЛИЗ КООРДИНАТ ЗАПОЛНЕНИЯ")
    print("=" * 80)
    
    # Поиск ключевых ячеек
    findings = {
        "document_number": None,
        "document_date": None,
        "implementation_date": None,
        "validity_period": None,
        "product": None,
        "reason": None,
        "tko_conclusion": None,
        "vruhcheno_marks": {},
        "table_start": None,
        "table_headers": {},
    }
    
    # Анализ шапки (строки 1-12)
    print("\nШАПКА ДОКУМЕНТА:")
    for row in range(1, 13):
        row_data = []
        for col in range(1, 16):
            cell = main_sheet.cell(row, col)
            if cell.value:
                val = str(cell.value).strip()
                row_data.append(f"{get_column_letter(col)}{row}: '{val[:40]}'")
        
        if row_data:
            print(f"  Строка {row}: {' | '.join(row_data[:5])}")
            
            # Поиск конкретных полей
            for col in range(1, 16):
                cell = main_sheet.cell(row, col)
                if cell.value:
                    val = str(cell.value).strip()
                    
                    # Номер документа
                    if "извещение" in val.lower() and "№" in val:
                        if findings["document_number"] is None:
                            # Ищем номер рядом
                            for c in range(col, min(col + 3, 16)):
                                num_cell = main_sheet.cell(row, c)
                                if num_cell.value and str(num_cell.value).strip().isdigit():
                                    findings["document_number"] = (row, c, num_cell.coordinate)
                                    print(f"    -> Номер документа: {num_cell.coordinate}")
                    
                    # Дата внедрения
                    if "дата внедрения" in val.lower():
                        # Дата обычно в следующей ячейке справа
                        date_cell = main_sheet.cell(row, col + 1)
                        if date_cell.value:
                            findings["implementation_date"] = (row, col + 1, date_cell.coordinate)
                            print(f"    -> Дата внедрения: {date_cell.coordinate}")
                    
                    # Срок действия
                    if "срок действия" in val.lower() or "партия" in val.lower():
                        period_cell = main_sheet.cell(row, col + 1)
                        if period_cell.value:
                            findings["validity_period"] = (row, col + 1, period_cell.coordinate)
                            print(f"    -> Срок действия: {period_cell.coordinate}")
                    
                    # Изделие
                    if "изделие" in val.lower() and ":" in val:
                        product_cell = main_sheet.cell(row, col + 1)
                        findings["product"] = (row, col + 1, product_cell.coordinate)
                        print(f"    -> Изделие: {product_cell.coordinate}")
                    
                    # Причина
                    if "причина" in val.lower() and ":" in val:
                        reason_cell = main_sheet.cell(row, col + 1)
                        findings["reason"] = (row, col + 1, reason_cell.coordinate)
                        print(f"    -> Причина: {reason_cell.coordinate}")
                    
                    # Заключение ТКО
                    if "заключение тко" in val.lower():
                        tko_cell = main_sheet.cell(row, col + 1)
                        findings["tko_conclusion"] = (row, col + 1, tko_cell.coordinate)
                        print(f"    -> Заключение ТКО: {tko_cell.coordinate}")
    
    # Анализ блока "Вручено"
    print("\nБЛОК 'ВРУЧЕНО':")
    workshops = ["ПЗУ", "ЗМУ", "СУ", "ОСУ", "ПДО", "ТКО", "ФЭО", "ОУП", "РСУ", "Кладовая ПДО"]
    
    for row in range(1, 15):
        for col in range(1, 16):
            cell = main_sheet.cell(row, col)
            if cell.value:
                val = str(cell.value).strip()
                
                # Ищем заголовки цехов
                for workshop in workshops:
                    if workshop in val:
                        # Ищем ячейку для отметки (обычно ниже или справа)
                        # Проверяем строку ниже
                        mark_cell_below = main_sheet.cell(row + 1, col)
                        if mark_cell_below.value and ("х" in str(mark_cell_below.value).lower() or "x" in str(mark_cell_below.value).lower()):
                            findings["vruhcheno_marks"][workshop] = (row + 1, col, mark_cell_below.coordinate)
                            print(f"  {workshop}: {mark_cell_below.coordinate}")
                        # Проверяем ячейку справа
                        mark_cell_right = main_sheet.cell(row, col + 1)
                        if mark_cell_right.value and ("х" in str(mark_cell_right.value).lower() or "x" in str(mark_cell_right.value).lower()):
                            findings["vruhcheno_marks"][workshop] = (row, col + 1, mark_cell_right.coordinate)
                            print(f"  {workshop}: {mark_cell_right.coordinate}")
    
    # Анализ таблицы изменений
    print("\nТАБЛИЦА ИЗМЕНЕНИЙ:")
    table_start_row = None
    
    for row in range(1, 30):
        row_text = " ".join([str(main_sheet.cell(row, c).value or "").strip() for c in range(1, 16)]).lower()
        
        if "подлежит изменению" in row_text or "вносимые изменения" in row_text:
            if table_start_row is None:
                table_start_row = row
                findings["table_start"] = row
                print(f"  Начало таблицы: строка {row}")
                
                # Анализ заголовков
                print("  Заголовки колонок:")
                for col in range(1, 16):
                    cell = main_sheet.cell(row, col)
                    if cell.value:
                        val = str(cell.value).strip()
                        if val:
                            findings["table_headers"][col] = (val, cell.coordinate)
                            print(f"    Колонка {col} ({cell.coordinate}): '{val[:50]}'")
                
                # Анализ следующей строки (подзаголовки)
                if row + 1 <= main_sheet.max_row:
                    print("  Подзаголовки:")
                    for col in range(1, 16):
                        cell = main_sheet.cell(row + 1, col)
                        if cell.value:
                            val = str(cell.value).strip()
                            if val and len(val) > 2:
                                print(f"    Колонка {col} ({cell.coordinate}): '{val[:50]}'")
    
    # Определяем область данных
    if table_start_row:
        data_start = table_start_row + 3  # предполагаем 2-3 строки заголовков
        print(f"\n  Область данных начинается со строки: {data_start}")
        
        # Показываем структуру первых строк данных
        print("  Примеры строк данных:")
        for row in range(data_start, min(data_start + 5, main_sheet.max_row + 1)):
            row_vals = []
            for col in range(1, 11):
                cell = main_sheet.cell(row, col)
                if cell.value:
                    val = str(cell.value).strip()
                    if val:
                        row_vals.append(f"{get_column_letter(col)}: '{val[:30]}'")
            if row_vals:
                print(f"    Строка {row}: {' | '.join(row_vals)}")
    
    # Определяем максимальное количество строк на первом листе
    # Обычно подписи начинаются с определённой строки
    print("\nОБЛАСТЬ ПОДПИСЕЙ:")
    signature_keywords = ["товаровед", "технолог", "конструктор", "начальник", "менеджер"]
    for row in range(40, min(60, main_sheet.max_row + 1)):
        row_text = " ".join([str(main_sheet.cell(row, c).value or "").strip() for c in range(1, 16)]).lower()
        for keyword in signature_keywords:
            if keyword in row_text:
                print(f"  Строка {row}: найдено '{keyword}'")
                break
    
    # Сохраняем результаты в файл конфигурации
    config = {
        "template_path": str(template_path),
        "main_sheet": main_sheet.title,
        "findings": {
            k: {"row": v[0], "col": v[1], "coord": v[2]} if v and isinstance(v, tuple) else v
            for k, v in findings.items()
        }
    }
    
    import json
    config_path = Path(__file__).parent.parent / "data" / "template_config.json"
    config_path.parent.mkdir(exist_ok=True, parents=True)
    
    with open(config_path, "w", encoding="utf-8") as f:
        json.dump(config, f, ensure_ascii=False, indent=2)
    
    print(f"\n[OK] Конфигурация сохранена: {config_path}")
    
    wb.close()

if __name__ == "__main__":
    detailed_analysis()

